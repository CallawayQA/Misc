import requests
from bs4 import BeautifulSoup
import openpyxl
import csv
import sys
sys.path.append('../')
from ToolLib import Library
from datetime import date



try:
    filePath = './/URL_List.xlsx'
    wb = openpyxl.load_workbook(filePath)
    sheet = wb['Sheet1']
    maxRowCount = sheet.max_row

except:
    print('######################################################')
    print('Please check the Excel file containing the list of URLs')
    print('######################################################')
    maxRowCount = -1


if maxRowCount > -1:
    for i in range(2, maxRowCount+1):
        try:
            response = requests.get(sheet.cell(i, 1).value)
            # print(response.status_code)
            if response.status_code == 200:

                print(sheet.cell(i, 1).value + ' is UP')
                sheet.cell(i, 3).value = 'UP'
                lLink = Library.listLink(sheet.cell(i, 1).value)
                today = date.today()
                fileName = sheet.cell(i, 2).value + '-' + today.strftime("%b-%d-%y")
                print(len(lLink))
                tmpList = lLink
                for j in range(len(lLink)):
                    if 'http' in lLink[j]:
                        print(lLink[j] + ' Outer Link')
                        print(Library.checkURL(lLink[j]))

                        Library.wrtText(fileName, lLink[j], Library.checkURL(lLink[j]))

                    else:
                        print(lLink[j] + " Inner Link")
                        print(sheet.cell(i, 1).value + lLink[j])
                        print(Library.checkURL(sheet.cell(i, 1).value + lLink[j]))
                        Library.wrtText(fileName, sheet.cell(i, 1).value + lLink[j], Library.checkURL(sheet.cell(i, 1).value + lLink[j]))
            else:
                print(sheet.cell(i, 1).value + 'is DOWN')
                sheet.cell(i, 3).value = 'DOWN'
        except Exception as e:
            print(sheet.cell(i, 1).value + ' is DOWN')
            sheet.cell(i, 3).value = 'DOWN'
            print(e)
            exception_type, exception_object, exception_traceback = sys.exc_info()
            filename = exception_traceback.tb_frame.f_code.co_filename
            line_number = exception_traceback.tb_lineno

            print("Exception type: ", exception_type)
            print("File name: ", filename)
            print("Line number: ", line_number)

    wb.save(filePath)
else:
    print('List is not available due to Input file issue')



