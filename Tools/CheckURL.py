import requests
import openpyxl

try:

    filePath = ".//iplist.xlsx"
    wb = openpyxl.load_workbook(filePath)
    sheet = wb['Sheet1']
    maxRowCount = sheet.max_row
except:
    print('######################################################')
    print('Please check the Excel file containing the list of IPs')
    print('######################################################')
    maxRowCount = -1

if maxRowCount > -1:
    for i in range(2, maxRowCount + 1):
        try:
            response = requests.get(sheet.cell(i, 1).value)
            # print(response.status_code)
            if response.status_code == 200:
                print(sheet.cell(i, 1).value + ' is UP')
                sheet.cell(i, 2).value = 'UP'
            else:
                print(sheet.cell(i, 1).value + 'is DOWN')
                sheet.cell(i, 2).value = 'DOWN'
        except:
            print(sheet.cell(i, 1).value + ' is DOWN')
            sheet.cell(i, 2).value = 'DOWN'

    wb.save(filePath)
else:
    print('List is not available due to Input file issue')